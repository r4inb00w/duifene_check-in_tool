import requests
import re
from flask import Flask, request, render_template, render_template_string
from openpyxl import load_workbook
import time


app = Flask(__name__)


@app.route('/', methods=['GET'])
def hello_world():
    return render_template('index.html')


@app.route('/index.html', methods=['GET'])
def hello_world1():
    return render_template('index.html')


@app.route('/index2.html', methods=['GET'])
def hello_world2():
    return render_template('index2.html')


@app.route('/index3.html', methods=['GET'])
def hello_world3():
    return render_template('index3.html')


@app.route('/index4.html', methods=['GET'])
def hello_world4():
    return render_template('index4.html')



@app.route('/code', methods=['POST'])
def code():
    code = request.form.get('code')
    if not code.isdigit():
        return '请输入纯数字！'
    code = code.replace(' ', '')
    if len(code) != 4:
        return '长度不正确，请重新输入！'
    time1 = time.time()
    result = check_in_code(code)
    time2 = time.time()
    times = round(time2 - time1, 2)
    # image = file("images/{}.jpg".format(7))
    return render_template_string(
        """{% for i in result %}
            <li>{{ i }}</li>
            {% endfor %}
            </br>
            总耗时为：{{times}}s
            <h3>点击<a href='/index.html>链接</a>返回</h3>
        """, result=result, times=times)

def console_excel():
    wb = load_workbook('users.xlsx')
    table = wb.active
    rows = table.max_row
    return table, rows


def write_excel_to_add_user(name, username, password, studentid):
    wb = load_workbook('users.xlsx')
    table = wb.active
    rows = table.max_row
    table.cell(rows + 1, 1).value = name
    table.cell(rows + 1, 2).value = username
    table.cell(rows + 1, 3).value = password
    table.cell(rows + 1, 4).value = studentid
    table.cell(rows + 1, 5).value = 'login'
    wb.save('users.xlsx')


@app.route('/add_user', methods=['POST'])
def add_user():
    username = request.form.get('username')
    password = request.form.get('password')
    name = request.form.get('name')
    studentid = request.form.get('studentid')
    if (not studentid.isdigit()) or (len(studentid) != 7):
        return 'studentid应为7个纯数字！'
    if username == '':
        return 'username不应该为空！'
    if password == '':
        return 'password不应该为空！'
    if name == '':
        return 'name不应该为空！'
    write_excel_to_add_user(name, username, password, studentid)
    return render_template_string(
        """
            <h1>成功写入该用户信息</h1>
            <h2>点击<a href='/index2.html'>链接</a>返回</h2>
        """)

@app.route('/select_user', methods=['POST'])
def show_users():
    name = request.form.get('name')
    studentid = ''
    flag = 0
    table, rows = console_excel()
    for i in range(1, rows + 1):
        c = table.cell(row=i, column=1).value
        if c == name:
            flag = 1
            studentid = table.cell(row=i, column=4).value
            break
    if flag != 1:
        temp = '没有找到该用户，请重新输入！'
    else:
        temp = 'name: ' + name + ', ' + 'studentid: ' + studentid
    return render_template_string(
        """
            <h1>{{ temp }}</h1>
            <h2>点击<a href='/index3.html'>链接</a>返回</h2>
        """, temp=temp)


@app.route('/delete_user', methods=['POST'])
def delete_user():
    name = request.form.get('name')
    username = request.form.get('username')
    password = request.form.get('password')
    flag = 0
    row = 0
    if username is None:
        return 'username不应该为空！'
    if password is None:
        return 'password不应该为空！'
    if name is None:
        return 'name不应该为空！'
    wb = load_workbook('users.xlsx')
    table = wb.active
    rows = table.max_row
    for i in range(1, rows + 1):
        c = table.cell(row=i, column=1).value
        if c == name:
            flag = 1
            row = i
            break
    if flag != 1:
        temp = '没有找到该用户，请重新输入想要删除的用户信息！'
    else:
        if(username == table.cell(row, 2).value) and (password == table.cell(row, 3).value):
            table.cell(row, 1).value = None
            table.cell(row, 2).value = None
            table.cell(row, 3).value = None
            table.cell(row, 4).value = None
            table.cell(row, 5).value = None
            wb.save('users.xlsx')
            temp = '删除成功！用户姓名为：' + name
        else:
            temp = 'username或password不正确，请重新输入'
    return render_template_string(
        """
            <h1>{{ temp }}</h1>
            <h2>点击<a href='/index4.html'>链接</a>返回</h2>
        """, temp=temp)


def todict():
    users = []
    table, rows = console_excel()
    for i in range(2, rows + 1):
        user = {'action': '',
                'loginname': '',
                'password': '',
                'studentid': '',
                'name': ''}  # Python for 循环中使用append()添加可变元素，前面的值会被覆盖，将user字典写入循环中即可
        user['action'] = table.cell(row=i, column=5).value
        user['loginname'] = table.cell(row=i, column=2).value
        user['password'] = table.cell(row=i, column=3).value
        user['studentid'] = table.cell(row=i, column=4).value
        user['name'] = table.cell(row=i, column=1).value
        if ((user['action'] is not None) and (user['loginname'] is not None) and (user['password'] is not None) and (
                user['studentid'] is not None) and (user['name'] is not None)):
            users.append(user)
    return users


def check_in_code(code):
    # 获取cookies
    result = []
    url = 'https://www.duifene.com/'
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0',
               'Connection': 'keep-alive'}
    req = requests.get(url=url, headers=headers)
    cookies = {'ASP.NET_SessionId': '',
               'tgw_l7_route': ''}
    regexp = []
    for i in req.cookies:
        pattern = re.compile('=\w+')
        regexp += re.findall(pattern, str(i))
    cookies['ASP.NET_SessionId'] = str(regexp[0]).replace('=', '')
    cookies['tgw_l7_route'] = str(regexp[1]).replace('=', '')

    # 获取用户信息
    users = todict()

    # 登录
    for data in users:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0',
                   'Accept': '*/*',
                   'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                   'Accept-Encoding': 'gzip, deflate',
                   'Referer': 'https://www.duifene.com/',
                   'Sec-Fetch-Dest': 'empty',
                   'Sec-Fetch-Mode': 'cors',
                   'Sec-Fetch-Site': 'same-origin'}

        url = 'https://www.duifene.com/AppCode/LoginInfo.ashx'
        # proxies={'http':'http://127.0.0.1:8080','https':'https://127.0.0.1:8080'}
        req = requests.post(url=url, headers=headers, data=data, cookies=cookies, allow_redirects=True)

        # 发送code
        url = 'https://www.duifene.com/_CheckIn/CheckIn.ashx'
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0',
                   'Connection': 'keep-alive'}
        files = {'action': (None, 'studentcheckin'),
                 'studentid': (None, data['studentid']),
                 'checkincode': (None, code)}
        req = requests.post(url=url, headers=headers, files=files, cookies=cookies, allow_redirects=True)
        result.append(data['name'] + ':' + req.text)

    return result


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=1012)
